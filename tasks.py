# tasks.py
import os
import zipfile
import pandas as pd
from sqlalchemy import text
from celery_app import celery_app
from csv_to_db.get_engine import get_engine
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import shutil

@celery_app.task(bind=True)
def generate_templates_task(self):
    # Connexion à la base et lecture de la table "test"
    engine = get_engine()
    engine.dispose()  
    engine = get_engine()
    df = pd.read_sql(text("SELECT * FROM fournisseur_produit WHERE status = 'ABSENT'"), engine)
    self.update_state(state='PROGRESS', meta={'progress': 10, 'message': 'Données chargées'})
    
    # Définition du mapping type/template
    type_produit_to_template = {
        'TLC': 'NEW Template_Produits_AGEC_Crystalchain_v3.xlsx',
        'ABJ': 'NEW Template_Produit_ABJ_Crystalchain.xlsx',
        'EEE': 'NEW Template_Produit_EEE_CrystalchainMonoprix.xlsx',
        'EMPAP': 'NEW Template_Produit_EMPAP_CrystalchainMonoprix.xlsx',
        'EA': 'NEW Template_Produit_MEUBLE_CrystalchainMonoprix.xlsx',
        'Jouet': 'NEW Template_Produit_JOUET_CrystalchainMonoprix.xlsx',
        'ASL': 'NEW Template_Produit_ASL_Crystalchain.xlsx'
    }
    
    column_mapping = {
        'entreprise': 'Entreprise',
        'nom_du_produit': 'NomProduit',
        'Pays_de_confection_ou_finition': 'PaysConfectionFinition',
        'ref_produits': 'Reference Monoprix',
        'node_name': 'Marque',
        'Réference_modele_pour_fournisseur': 'ReferenceProduit',
        'URL_photo': 'URLPhotoProduit',
        'Présence_Substances_Dangereuses': 'SubstancesDangereuses',
        'Emballage__recyclabilité': 'EmballageRecyclabilite',
        'Emballage__présence_substances_dangereuses': 'EmballageSubstancesDangereuses',
        'Collection': 'Collection',
        'Description': 'Description',
        'Informations_supplémentaires_fournisseur': 'InformationsSupplementaires',
        'Commentaires': 'Commentaires',
        "EAN_UVC": "EAN"
    }
    
    # Appliquer les filtres
    if 'status' in df.columns:
        df = df[df['status'] == 'ABSENT']
        print("status existe !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    df = df[df['Type_de_produit'].str.contains('|'.join(type_produit_to_template.keys()), case=False, na=False)]
    print(df['status'])
    # Regroupement par fournisseur (en supposant que la colonne "nom_fournisseur" existe)
    groups = list(df.groupby('nom_fournisseur'))
    total_groups = len(groups)
    current_group = 0
    
    # Utiliser un répertoire partagé pour stocker les fichiers générés
    shared_dir = "/tmp/shared"  
    # Vérifier si le dossier existe
    if os.path.exists(shared_dir):
        # Supprimer uniquement les fichiers et sous-dossiers sans supprimer `shared_dir`
        for filename in os.listdir(shared_dir):
            file_path = os.path.join(shared_dir, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path) 
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path) 
            except Exception as e:
                print(f"Erreur lors de la suppression de {file_path}: {e}")

    # S'assurer que le dossier existe bien après le nettoyage
    os.makedirs(shared_dir, exist_ok=True)
    zip_filename = os.path.join(shared_dir, 'generated_templates.zip')
    
    # Répertoire contenant les templates (à adapter selon votre environnement)
    templates_dir = 'NEW Templates'
    
    for fournisseur, data in groups:
        # print(f"Fournisseur: {fournisseur}, Status uniques: {data['status'].unique()}")

        current_group += 1
        progress = 10 + int((current_group / total_groups) * 80)  
        self.update_state(state='PROGRESS', meta={'progress': progress, 'message': f'Traitement du fournisseur {fournisseur}'})
        
        fournisseur_cleaned = fournisseur.replace("/", "-").replace("\\", "-")
        fournisseur_dir = os.path.join(shared_dir, fournisseur_cleaned)
        # print("fournisseur_dir: ", fournisseur_dir)
        os.makedirs(fournisseur_dir, exist_ok=True)
        
        for type_produit, template_file_name in type_produit_to_template.items():
            fournisseur_data = data[data['Type_de_produit'].str.contains(type_produit, case=False, na=False)]
            # print(f"Fournisseur: {fournisseur}, Type: {type_produit}, Status: {fournisseur_data['status'].unique()}")
            # print("fournisseur data : ", fournisseur_data)
            if fournisseur_data.empty:
                continue
            
            template_file = os.path.join(templates_dir, template_file_name)
            if not os.path.exists(template_file):
                continue
            
            wb = load_workbook(template_file)
            if len(wb.sheetnames) < 2:
                continue
            ws = wb.worksheets[1]
            start_row = 8
            
            # Mettre en forme l'en-tête
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF")
            
            if "Format" not in wb.sheetnames:
                continue
            format_ws = wb["Format"]
            for row in format_ws.iter_rows(min_row=2, max_row=format_ws.max_row, min_col=1, max_col=18):
                field_name = row[0].value
                field_type = row[2].value
                if field_type == 'enum':
                    dropdown_values = [cell.value for cell in row[5:18] if cell.value]
                    if dropdown_values:
                        values_string = ','.join(dropdown_values)
                        template_col_letter = None
                        for cell in ws[1]:
                            if cell.value == field_name:
                                template_col_letter = cell.column_letter
                                break
                        if template_col_letter:
                            dv = DataValidation(
                                type="list",
                                formula1=f'"{values_string}"',
                                showDropDown=False
                            )
                            ws.add_data_validation(dv)
                            dv.add(f'{template_col_letter}{start_row}:{template_col_letter}1048576')
            
            # Remplissage des colonnes d'après le mapping
            for input_col, template_col in column_mapping.items():
                col_letter = None
                for cell in ws[1]:
                    if cell.value == template_col:
                        col_letter = cell.column_letter
                        break
                if col_letter:
                    for idx, value in enumerate(fournisseur_data[input_col].values, start=start_row):
                        ws[f'{col_letter}{idx}'].value = value
            
            # Ajout (ou récupération) de la colonne "Type_de_produit"
            if 'Type_de_produit' not in [cell.value for cell in ws[1]]:
                new_col_index = ws.max_column + 1
                new_col_letter = get_column_letter(new_col_index)
                ws[f'{new_col_letter}1'].value = 'Type_de_produit'
                ws[f'{new_col_letter}1'].font = Font(color="000000")
            else:
                new_col_letter = [cell.column_letter for cell in ws[1] if cell.value == 'Type_de_produit'][0]
            
            # Remplissage des colonnes TraceNumber et TraceType
            trace_number_col = None
            trace_type_col = None
            for cell in ws[1]:
                if cell.value == 'TraceNumber':
                    trace_number_col = cell.column_letter
                elif cell.value == 'TraceType':
                    trace_type_col = cell.column_letter
            if trace_number_col and trace_type_col:
                for idx in range(start_row, len(fournisseur_data) + start_row):
                    ws[f'{trace_number_col}{idx}'].value = idx - (start_row - 1)
                    if type_produit == 'TLC':
                        ws[f'{trace_type_col}{idx}'].value = 'PRODUIT-AGEC'
                    elif type_produit == "EA":
                        ws[f'{trace_type_col}{idx}'].value = 'PRODUIT-Meuble'
                    else:
                        ws[f'{trace_type_col}{idx}'].value = f'PRODUIT-{type_produit}'
            
            # Gestion de la colonne "Deactivated"
            deactivated_col_letter = None
            for cell in ws[1]:
                if cell.value == 'Deactivated':
                    deactivated_col_letter = cell.column_letter
                    break
            if deactivated_col_letter:
                for idx in range(start_row, len(fournisseur_data) + start_row):
                    row_has_data = any(ws[f'{c.column_letter}{idx}'].value for c in ws[1])
                    if row_has_data:
                        ws[f'{deactivated_col_letter}{idx}'].value = 'false'
            
            # Suppression des doublons sur "Reference Monoprix"
            reference_monoprix_col = None
            for cell in ws[1]:
                if cell.value == 'Reference Monoprix':
                    reference_monoprix_col = cell.column_letter
                    break
            if reference_monoprix_col:
                unique_values = set()
                rows_to_delete = []
                for row in range(start_row, ws.max_row + 1):
                    cell_value = ws[f'{reference_monoprix_col}{row}'].value
                    if cell_value in unique_values:
                        rows_to_delete.append(row)
                    else:
                        unique_values.add(cell_value)
                for row in reversed(rows_to_delete):
                    ws.delete_rows(row)
            
            # Ajustement des commentaires
            for cell in ws[2]:
                cell.font = Font(color="FFFFFF")
                if cell.comment:
                    comment_text = cell.comment.text
                    cell.comment = None
                    new_comment = Comment(comment_text, "Auteur")
                    cell.comment = new_comment
                    cell.comment.width = 400
                    cell.comment.height = 300
            
            output_file = os.path.join(fournisseur_dir, f'{fournisseur_cleaned}_{type_produit}_CrystalChainMonoprix.xlsx')
            wb.save(output_file)
        
    self.update_state(state='PROGRESS', meta={'progress': 90, 'message': 'Création du fichier ZIP'})
    with zipfile.ZipFile(zip_filename, 'w') as zf:
        for root, dirs, files in os.walk(shared_dir):
            for file in files:
                if file != 'generated_templates.zip':
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, shared_dir)
                    # print(f"Adding file {file_path} as {arcname}")
                    zf.write(file_path, arcname)
    
    # self.update_state(state='SUCCESS', meta={'progress': 100, 'message': 'Génération terminée'})
    final_result = str(zip_filename)
    # print("Returning file path:", final_result)
    return final_result
